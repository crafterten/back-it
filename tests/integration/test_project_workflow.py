import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from backit.projects import (
    check_project,
    configure_project,
    ingest_project,
    init_project,
    inspect_project,
    load_project_config,
)


def _write_scores(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scores"
    sheet.append(
        [
            "observed_at",
            "score_received_at",
            "study_hours",
            "sleep_hours",
            "score",
        ]
    )
    sheet.append(
        [
            "2026-07-01T09:00:00-04:00",
            "2026-07-02T15:00:00-04:00",
            2.5,
            7.0,
            82,
        ]
    )
    workbook.save(path)


def _config() -> dict:
    return {
        "schema_name": "backit.project",
        "schema_version": 1,
        "project_id": "test-scores",
        "name": "Test Scores",
        "goal": "Predict my next test score.",
        "state": "configured",
        "canonical_workbook": "source/project_data.xlsx",
        "sheet": "Scores",
        "timezone": "America/New_York",
        "columns": [
            {
                "name": "observed_at",
                "semantic_type": "datetime",
                "required": True,
                "available_at": {"same_as": "observed_at"},
            },
            {
                "name": "score_received_at",
                "semantic_type": "datetime",
                "required": True,
                "available_at": {"same_as": "observed_at"},
            },
            {
                "name": "study_hours",
                "semantic_type": "number",
                "unit": "hours",
                "required": True,
                "minimum": 0,
                "available_at": {"same_as": "observed_at"},
            },
            {
                "name": "sleep_hours",
                "semantic_type": "number",
                "unit": "hours",
                "required": True,
                "minimum": 0,
                "maximum": 24,
                "available_at": {"same_as": "observed_at"},
            },
        ],
        "outcome": {
            "column": "score",
            "semantic_type": "score",
            "unit": "points",
            "better": "higher",
            "minimum": 0,
            "maximum": 100,
            "available_at": {"column": "score_received_at"},
        },
        "adapter": {"name": "test_scores", "version": "1"},
    }


def test_project_lifecycle_creates_immutable_snapshot(tmp_path: Path) -> None:
    source = tmp_path / "loose.xlsx"
    project = tmp_path / "test-scores"
    _write_scores(source)
    source_bytes = source.read_bytes()

    initialized = init_project(
        project,
        source,
        name="Test Scores",
        goal="Predict my next test score.",
        idempotency_key="init-test-scores",
    )
    assert initialized.state == "unconfigured"
    assert source.read_bytes() == source_bytes

    inspection = inspect_project(project)
    assert inspection.state == "inspected"
    assert inspection.headers[-1] == "score"

    configured = configure_project(
        project,
        _config(),
        idempotency_key="configure-test-scores-v1",
    )
    assert configured.state == "configured"
    assert load_project_config(project).project_id == "test-scores"

    canonical = project / "source" / "project_data.xlsx"
    workbook = load_workbook(canonical)
    sheet = workbook["Scores"]
    sheet.append(
        [
            None,
            "2026-07-08T09:00:00-04:00",
            "2026-07-09T15:00:00-04:00",
            4.0,
            8.0,
            None,
        ]
    )
    workbook.save(canonical)

    ingested = ingest_project(project)

    assert ingested.row_count == 2
    assert ingested.snapshot_path.exists()
    assert ingested.manifest_path.exists()
    manifest = json.loads(ingested.manifest_path.read_text(encoding="utf-8"))
    assert len(manifest["environment"]["source_tree_hash"]) == 64
    assert manifest["environment"]["python"].startswith("3.")
    workbook = load_workbook(canonical)
    assert workbook["Scores"].cell(3, 1).value

    workbook = load_workbook(canonical)
    sheet = workbook["Scores"]
    sheet.append(
        [
            None,
            "2026-07-15T09:00:00-04:00",
            "2026-07-16T15:00:00-04:00",
            5.0,
            7.5,
            94,
        ]
    )
    workbook.save(canonical)
    before_check = canonical.read_bytes()

    checked = check_project(project)

    assert checked.ok is True
    assert checked.new_rows == 1
    assert canonical.read_bytes() == before_check

    workbook = load_workbook(canonical)
    workbook["Scores"].cell(4, 4).value = "five-ish"
    workbook.save(canonical)
    invalid_bytes = canonical.read_bytes()

    invalid = check_project(project)

    assert invalid.ok is False
    assert invalid.diagnostics[0]["column"] == "study_hours"
    assert invalid.diagnostics[0]["excel_row"] == 4
    assert canonical.read_bytes() == invalid_bytes


def test_init_hashes_idempotency_key_before_using_it_as_a_filename(
    tmp_path: Path,
) -> None:
    source = tmp_path / "loose.xlsx"
    project = tmp_path / "safe-project"
    _write_scores(source)

    result = init_project(
        project,
        source,
        name="Safe Project",
        goal="Keep idempotency keys out of paths.",
        idempotency_key="../unsafe:key",
    )

    assert result.state == "unconfigured"
    assert len(list((project / "audit").glob("idempotency-*.json"))) == 1
