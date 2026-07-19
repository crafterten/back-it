from pathlib import Path

from openpyxl import Workbook, load_workbook

from backit.workbook import create_canonical_workbook, inspect_workbook


def _write_scores(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Scores"
    sheet.append(["observed_at", "study_hours", "sleep_hours", "score"])
    sheet.append(["2026-07-01T20:00:00-04:00", 2.5, 7.0, 82])
    sheet.append(["2026-07-08T20:00:00-04:00", 4.0, 8.0, 91])
    workbook.save(path)


def test_inspect_and_create_canonical_workbook(tmp_path: Path) -> None:
    source = tmp_path / "loose.xlsx"
    canonical = tmp_path / "project_data.xlsx"
    _write_scores(source)
    before = source.read_bytes()

    inspection = inspect_workbook(source)
    assert inspection.sheet == "Scores"
    assert inspection.headers == [
        "observed_at",
        "study_hours",
        "sleep_hours",
        "score",
    ]
    assert inspection.row_count == 2
    assert inspection.errors == []

    created = create_canonical_workbook(
        source=source,
        destination=canonical,
        sheet_name="Scores",
    )

    assert source.read_bytes() == before
    assert created.rows_with_ids == 2
    workbook = load_workbook(canonical, data_only=False)
    sheet = workbook["Scores"]
    assert sheet.cell(1, 1).value == "_backit_row_id"
    first_id = sheet.cell(2, 1).value
    second_id = sheet.cell(3, 1).value
    assert isinstance(first_id, str)
    assert first_id != second_id
