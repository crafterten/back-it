from dataclasses import dataclass
import os
from pathlib import Path
import tempfile
from uuid import uuid4

from openpyxl import Workbook, load_workbook


@dataclass(frozen=True)
class WorkbookInspection:
    sheet: str
    headers: list[str]
    row_count: int
    errors: list[str]
    warnings: list[str]


@dataclass(frozen=True)
class CanonicalWorkbookResult:
    path: Path
    rows_with_ids: int


def inspect_workbook(path: Path) -> WorkbookInspection:
    workbook = load_workbook(path, data_only=False, read_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    errors: list[str] = []
    warnings: list[str] = []

    if sheet.merged_cells.ranges:
        errors.append("MERGED_CELLS_UNSUPPORTED")

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in sheet[1]
    ]
    if not headers or any(not header for header in headers):
        errors.append("HEADER_REQUIRED")
    if len(headers) != len(set(headers)):
        errors.append("DUPLICATE_HEADERS")

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.data_type == "f":
                errors.append(f"FORMULA_UNSUPPORTED:{cell.coordinate}")
    for index, dimension in sheet.row_dimensions.items():
        if index > 1 and dimension.hidden:
            warnings.append(f"HIDDEN_ROW:{index}")
    for name, dimension in sheet.column_dimensions.items():
        if dimension.hidden:
            warnings.append(f"HIDDEN_COLUMN:{name}")

    row_count = sum(
        1
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if any(value is not None for value in row)
    )
    return WorkbookInspection(
        sheet=sheet.title,
        headers=headers,
        row_count=row_count,
        errors=sorted(set(errors)),
        warnings=sorted(set(warnings)),
    )


def create_canonical_workbook(
    source: Path,
    destination: Path,
    sheet_name: str,
) -> CanonicalWorkbookResult:
    inspection = inspect_workbook(source)
    if inspection.errors:
        raise ValueError(";".join(inspection.errors))

    source_workbook = load_workbook(source, data_only=False, read_only=False)
    source_sheet = source_workbook[sheet_name]
    target_workbook = Workbook()
    target_sheet = target_workbook.active
    target_sheet.title = sheet_name
    target_sheet.append(["_backit_row_id", *inspection.headers])

    rows_with_ids = 0
    for values in source_sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in values):
            continue
        target_sheet.append([str(uuid4()), *values])
        rows_with_ids += 1

    destination.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(
        suffix=".xlsx",
        prefix=f".{destination.stem}-",
        dir=destination.parent,
    )
    os.close(handle)
    temporary = Path(temporary_name)
    try:
        target_workbook.save(temporary)
        os.replace(temporary, destination)
    finally:
        temporary.unlink(missing_ok=True)
    return CanonicalWorkbookResult(destination, rows_with_ids)
