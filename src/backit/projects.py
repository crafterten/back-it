from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import json
import os
from pathlib import Path
import platform
import shutil
import sys
import tempfile
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook
import pandas as pd
import portalocker
import yaml

from backit.contracts import ColumnConfig, ProjectConfig
from backit.provenance import source_tree_hash
from backit.serialization import canonical_json, sha256_json
from backit.workbook import create_canonical_workbook, inspect_workbook


PROJECT_DIRS = (
    "audit",
    "config",
    "ledger",
    "plans",
    "reports",
    "snapshots",
    "source",
    "transactions",
)


@dataclass(frozen=True)
class ProjectStateResult:
    state: str


@dataclass(frozen=True)
class ProjectInspectionResult:
    state: str
    sheet: str
    headers: list[str]
    row_count: int
    errors: list[str]
    warnings: list[str]


@dataclass(frozen=True)
class IngestResult:
    snapshot_id: str
    row_count: int
    snapshot_path: Path
    manifest_path: Path


@dataclass(frozen=True)
class CheckResult:
    ok: bool
    new_rows: int
    changed_rows: int
    deleted_rows: int
    diagnostics: list[dict[str, Any]]


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def _atomic_bytes(path: Path, payload: bytes) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary_name = tempfile.mkstemp(
        prefix=f".{path.name}-",
        dir=path.parent,
    )
    try:
        with os.fdopen(handle, "wb") as stream:
            stream.write(payload)
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary_name, path)
    finally:
        Path(temporary_name).unlink(missing_ok=True)


def _write_json(path: Path, value: Any) -> None:
    _atomic_bytes(path, canonical_json(value))


def _write_yaml(path: Path, value: Any) -> None:
    payload = yaml.safe_dump(
        value,
        allow_unicode=True,
        sort_keys=True,
    ).encode("utf-8")
    _atomic_bytes(path, payload)


def _read_yaml(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as stream:
        value = yaml.safe_load(stream)
    if not isinstance(value, dict):
        raise ValueError(f"Expected object in {path}")
    return value


def _original_import(project: Path) -> Path:
    matches = sorted((project / "source").glob("original-import-*.xlsx"))
    if len(matches) != 1:
        raise ValueError("Project must contain exactly one original import")
    return matches[0]


def _idempotency_path(project: Path, key: str) -> Path:
    digest = hashlib.sha256(key.encode("utf-8")).hexdigest()
    return project / "audit" / f"idempotency-{digest}.json"


def init_project(
    project: Path,
    workbook: Path,
    *,
    name: str,
    goal: str,
    idempotency_key: str,
) -> ProjectStateResult:
    if not idempotency_key:
        raise ValueError("idempotency_key is required")
    if workbook.suffix.lower() != ".xlsx":
        raise ValueError("MVP supports only .xlsx workbooks")

    for directory in PROJECT_DIRS:
        (project / directory).mkdir(parents=True, exist_ok=True)

    idempotency_path = _idempotency_path(project, idempotency_key)
    if idempotency_path.exists():
        state = _read_yaml(project / "project.yaml")["state"]
        return ProjectStateResult(state)

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    copied = project / "source" / f"original-import-{timestamp}.xlsx"
    shutil.copy2(workbook, copied)
    state = {
        "project_id": project.name,
        "name": name,
        "goal": goal,
        "state": "unconfigured",
        "original_import": copied.relative_to(project).as_posix(),
        "created_at": _utc_now(),
    }
    _write_yaml(project / "project.yaml", state)
    _write_json(
        idempotency_path,
        {"key": idempotency_key, "state": "unconfigured", "created_at": _utc_now()},
    )
    return ProjectStateResult("unconfigured")


def inspect_project(project: Path) -> ProjectInspectionResult:
    inspection = inspect_workbook(_original_import(project))
    plan = {
        "schema_name": "backit.import-plan",
        "schema_version": 1,
        "sheet": inspection.sheet,
        "headers": inspection.headers,
        "row_count": inspection.row_count,
        "errors": inspection.errors,
        "warnings": inspection.warnings,
    }
    _write_json(project / "plans" / "import-plan.json", plan)
    state = _read_yaml(project / "project.yaml")
    state["state"] = "inspected"
    _write_yaml(project / "project.yaml", state)
    return ProjectInspectionResult(
        state="inspected",
        sheet=inspection.sheet,
        headers=inspection.headers,
        row_count=inspection.row_count,
        errors=inspection.errors,
        warnings=inspection.warnings,
    )


def configure_project(
    project: Path,
    raw_config: dict[str, Any],
    *,
    idempotency_key: str,
) -> ProjectStateResult:
    if not idempotency_key:
        raise ValueError("idempotency_key is required")
    config = ProjectConfig.model_validate(raw_config)
    plan_path = project / "plans" / "import-plan.json"
    if not plan_path.exists():
        raise ValueError("Project must be inspected before configuration")
    plan = json.loads(plan_path.read_text(encoding="utf-8"))
    if plan["errors"]:
        raise ValueError("Import plan contains blocking errors")

    destination = project / config.canonical_workbook
    create_canonical_workbook(
        source=_original_import(project),
        destination=destination,
        sheet_name=config.sheet,
    )
    config_data = config.model_dump(mode="json")
    config_hash = sha256_json(config_data)
    immutable = (
        project
        / "config"
        / f"schema-v{config.schema_version}-{config_hash[:12]}.yaml"
    )
    _write_yaml(immutable, config_data)
    _write_yaml(project / "project.yaml", config_data)
    _write_json(
        _idempotency_path(project, idempotency_key),
        {
            "key": idempotency_key,
            "state": "configured",
            "config_hash": config_hash,
        },
    )
    return ProjectStateResult("configured")


def load_project_config(project: Path) -> ProjectConfig:
    return ProjectConfig.model_validate(_read_yaml(project / "project.yaml"))


def _coerce_cell(value: Any, column: ColumnConfig) -> Any:
    if value is None:
        if column.required:
            raise ValueError(f"{column.name}: value is required")
        return None
    if column.semantic_type in {"duration", "pace"}:
        if isinstance(value, bool):
            raise ValueError(f"{column.name}: expected duration")
        if isinstance(value, (int, float)):
            number = float(value)
        elif isinstance(value, str):
            parts = value.strip().split(":")
            if len(parts) not in {2, 3}:
                raise ValueError(f"{column.name}: expected MM:SS or HH:MM:SS")
            try:
                parsed = [float(part) for part in parts]
            except ValueError as error:
                raise ValueError(
                    f"{column.name}: expected MM:SS or HH:MM:SS"
                ) from error
            if parsed[-1] < 0 or parsed[-1] >= 60:
                raise ValueError(f"{column.name}: seconds must be under 60")
            if len(parts) == 2:
                number = parsed[0] * 60 + parsed[1]
            else:
                if parsed[1] < 0 or parsed[1] >= 60:
                    raise ValueError(f"{column.name}: minutes must be under 60")
                number = parsed[0] * 3600 + parsed[1] * 60 + parsed[2]
        else:
            raise ValueError(f"{column.name}: expected duration")
        if column.minimum is not None and number < column.minimum:
            raise ValueError(f"{column.name}: below minimum")
        if column.maximum is not None and number > column.maximum:
            raise ValueError(f"{column.name}: above maximum")
        return number
    if column.semantic_type in {
        "number",
        "score",
        "distance",
        "percentage",
    }:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValueError(f"{column.name}: expected number")
        number = float(value)
        if column.minimum is not None and number < column.minimum:
            raise ValueError(f"{column.name}: below minimum")
        if column.maximum is not None and number > column.maximum:
            raise ValueError(f"{column.name}: above maximum")
        return number
    if column.semantic_type == "integer":
        if isinstance(value, bool) or not isinstance(value, int):
            raise ValueError(f"{column.name}: expected integer")
        return value
    if column.semantic_type == "datetime":
        parsed = pd.to_datetime(value, utc=True, errors="raise")
        return parsed.isoformat()
    return str(value)


def _schema_columns(config: ProjectConfig) -> list[ColumnConfig]:
    return [
        *config.columns,
        ColumnConfig(
            name=config.outcome.column,
            semantic_type=config.outcome.semantic_type,
            unit=config.outcome.unit,
            required=False,
            minimum=config.outcome.minimum,
            maximum=config.outcome.maximum,
            available_at=config.outcome.available_at,
        ),
    ]


def _latest_manifest(project: Path) -> dict[str, Any] | None:
    manifests: list[dict[str, Any]] = []
    for path in (project / "snapshots").glob("*-manifest.json"):
        manifests.append(json.loads(path.read_text(encoding="utf-8")))
    return max(manifests, key=lambda value: value["created_at"]) if manifests else None


def check_project(project: Path) -> CheckResult:
    config = load_project_config(project)
    workbook_path = project / config.canonical_workbook
    workbook = load_workbook(workbook_path, data_only=False)
    sheet = workbook[config.sheet]
    headers = [cell.value for cell in sheet[1]]
    schema_columns = _schema_columns(config)
    expected = ["_backit_row_id", *(column.name for column in schema_columns)]
    diagnostics: list[dict[str, Any]] = []
    if headers != expected:
        diagnostics.append(
            {
                "code": "HEADERS_MISMATCH",
                "message": f"Expected {expected!r}, found {headers!r}",
                "excel_row": 1,
                "column": None,
            }
        )
        return CheckResult(False, 0, 0, 0, diagnostics)

    latest = _latest_manifest(project) or {"row_hashes": {}}
    prior_hashes: dict[str, str] = latest.get("row_hashes", {})
    current_hashes: dict[str, str] = {}
    seen_ids: set[str] = set()
    new_rows = 0
    for excel_row, values in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        if not any(value is not None for value in values):
            continue
        row_id = values[0]
        if row_id is None:
            row_id = f"pending:{excel_row}"
            new_rows += 1
        elif not isinstance(row_id, str) or not row_id:
            diagnostics.append(
                {
                    "code": "ROW_ID_INVALID",
                    "message": "Expected a non-empty string row ID",
                    "excel_row": excel_row,
                    "column": "_backit_row_id",
                }
            )
            continue
        elif row_id in seen_ids:
            diagnostics.append(
                {
                    "code": "ROW_ID_DUPLICATE",
                    "message": "Duplicate row ID",
                    "excel_row": excel_row,
                    "column": "_backit_row_id",
                }
            )
            continue
        else:
            seen_ids.add(row_id)
            if row_id not in prior_hashes:
                new_rows += 1

        record: dict[str, Any] = {"_backit_row_id": row_id}
        for column, value in zip(schema_columns, values[1:], strict=True):
            try:
                record[column.name] = _coerce_cell(value, column)
            except (TypeError, ValueError) as error:
                diagnostics.append(
                    {
                        "code": "CELL_INVALID",
                        "message": str(error),
                        "excel_row": excel_row,
                        "row_id": row_id,
                        "column": column.name,
                        "raw_value": value,
                    }
                )
        if not any(item.get("excel_row") == excel_row for item in diagnostics):
            current_hashes[row_id] = sha256_json(record)

    changed_rows = sum(
        1
        for row_id, row_hash in current_hashes.items()
        if row_id in prior_hashes and prior_hashes[row_id] != row_hash
    )
    deleted_rows = len(set(prior_hashes) - set(current_hashes))
    return CheckResult(
        ok=not diagnostics,
        new_rows=new_rows,
        changed_rows=changed_rows,
        deleted_rows=deleted_rows,
        diagnostics=diagnostics,
    )


def ingest_project(project: Path) -> IngestResult:
    config = load_project_config(project)
    workbook_path = project / config.canonical_workbook
    lock_path = project / ".backit.lock"
    with portalocker.Lock(lock_path, timeout=0):
        workbook = load_workbook(workbook_path, data_only=False)
        sheet = workbook[config.sheet]
        headers = [cell.value for cell in sheet[1]]
        expected = [
            "_backit_row_id",
            *(column.name for column in config.columns),
            config.outcome.column,
        ]
        if headers != expected:
            raise ValueError(f"Workbook headers do not match schema: {headers!r}")

        schema_columns = _schema_columns(config)
        seen_ids: set[str] = set()
        rows: list[dict[str, Any]] = []
        assigned_ids = False
        for excel_row, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            if not any(value is not None for value in values):
                continue
            row_id = values[0]
            if row_id is None:
                row_id = str(uuid4())
                sheet.cell(excel_row, 1).value = row_id
                assigned_ids = True
            if not isinstance(row_id, str) or not row_id:
                raise ValueError(f"Row {excel_row}: invalid _backit_row_id")
            if row_id in seen_ids:
                raise ValueError(f"Row {excel_row}: duplicate _backit_row_id")
            seen_ids.add(row_id)
            record: dict[str, Any] = {"_backit_row_id": row_id}
            for column, value in zip(schema_columns, values[1:], strict=True):
                record[column.name] = _coerce_cell(value, column)
            rows.append(record)

        typed_rows_hash = sha256_json(rows)
        snapshot_id = sha256_json(
            {
                "typed_rows_hash": typed_rows_hash,
                "schema_version": config.schema_version,
                "adapter": config.adapter.model_dump(mode="json"),
            }
        )
        snapshot_path = project / "snapshots" / f"{snapshot_id}.parquet"
        manifest_path = project / "snapshots" / f"{snapshot_id}-manifest.json"

        if assigned_ids:
            handle, temporary_name = tempfile.mkstemp(
                suffix=".xlsx",
                prefix=".project-data-",
                dir=workbook_path.parent,
            )
            os.close(handle)
            temporary = Path(temporary_name)
            try:
                workbook.save(temporary)
                os.replace(temporary, workbook_path)
            finally:
                temporary.unlink(missing_ok=True)

        if not snapshot_path.exists():
            frame = pd.DataFrame(rows)
            handle, temporary_name = tempfile.mkstemp(
                suffix=".parquet",
                prefix=".snapshot-",
                dir=snapshot_path.parent,
            )
            os.close(handle)
            temporary = Path(temporary_name)
            try:
                frame.to_parquet(temporary, index=False)
                os.replace(temporary, snapshot_path)
            finally:
                temporary.unlink(missing_ok=True)

        manifest = {
            "schema_name": "backit.snapshot-manifest",
            "schema_version": 1,
            "snapshot_id": snapshot_id,
            "created_at": _utc_now(),
            "typed_rows_hash": typed_rows_hash,
            "row_count": len(rows),
            "row_ids": sorted(seen_ids),
            "row_hashes": {
                str(row["_backit_row_id"]): sha256_json(row)
                for row in rows
            },
            "project_schema_version": config.schema_version,
            "adapter": config.adapter.model_dump(mode="json"),
            "environment": {
                "python": platform.python_version(),
                "platform": platform.platform(),
                "implementation": sys.implementation.name,
                "source_tree_hash": source_tree_hash(
                    Path(__file__).resolve().parents[2]
                ),
                "lock_hash": hashlib.sha256(
                    (
                        Path(__file__).resolve().parents[2]
                        / "requirements.lock"
                    ).read_bytes()
                ).hexdigest(),
            },
            "files": {"data": snapshot_path.name},
        }
        _write_json(manifest_path, manifest)
        _write_yaml(
            project / "project.yaml",
            config.model_copy(update={"state": "ingested"}).model_dump(mode="json"),
        )
        return IngestResult(
            snapshot_id=snapshot_id,
            row_count=len(rows),
            snapshot_path=snapshot_path,
            manifest_path=manifest_path,
        )
